import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]



cell = sheet["a1"]
column = sheet["a1"]
column = sheet["a"]
cells = column["a:c"]
sheet["a1:c3"]
sheet[1:3]
print(cells)

print(column)
print(cell.row)
print(cell.column)
print(cell.coordinate)

print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
    for column in range(row, column):
        cell = sheet.cell(row, column)
        print(cell.value)

sheet.append([1,2,3])
wb.save("transactions2.xlsx")